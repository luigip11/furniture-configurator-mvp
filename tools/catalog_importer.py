#!/usr/bin/env python3
"""Genera audit e manifest per importare il catalogo ricevuto dal committente.

Lo script legge l'Excel di staging, incrocia i GLB disponibili e produce file
intermedi verificabili prima di caricare dati e asset su Supabase.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INCOMING_DIR = ROOT / "incoming" / "catalogo-committente"
DEFAULT_OUTPUT_DIR = DEFAULT_INCOMING_DIR / "audit"
ASSET_BUCKET = "product-assets"
MAX_MODEL_SIZE_MB = 80


@dataclass(frozen=True)
class Dimensions:
    width_mm: int
    height_mm: int
    depth_mm: int


@dataclass(frozen=True)
class CatalogProduct:
    code: str
    name_it: str
    name_en: str | None
    category_slug: str
    category_name: str
    width_mm: int
    height_mm: int
    depth_mm: int
    thickness_mm: float | None
    price: float | None
    preview_image_url: str | None
    model_url: str | None
    technical_file_url: str | None
    is_published: bool
    local_model_path: str
    storage_model_path: str
    variant_key: str
    excel_match: str | None
    review_flags: list[str]


CATEGORY_LABELS = {
    "basi": "Basi",
    "basi-lavatrice-asciugatrice": "Basi lavatrice-asciugatrice",
    "colonne": "Colonne",
    "colonne-impianti": "Colonne impianti",
    "contenitori-impianti": "Contenitori impianti",
    "pensili": "Pensili",
    "portali": "Portali",
}


# Mantiene esplicita la corrispondenza fra nomi GLB reali e famiglie tecniche.
FAMILY_RULES = [
    (
        "pensile-verticale",
        ["pensile verticale"],
        "pensili",
        Dimensions(700, 878, 350),
        "PENSILE VERTICALE",
    ),
    (
        "pensile-orizzontale",
        ["pensile orizzontale"],
        "pensili",
        Dimensions(700, 439, 350),
        "PENSILE ORIZZONTALE",
    ),
    (
        "colonna-mista-impianto-alto",
        ["colonna mista impianto alto"],
        "colonne-impianti",
        Dimensions(700, 2282, 665),
        "COLONNA CON",
    ),
    (
        "colonna-mista-impianto-basso",
        ["colonna mista impianto basso"],
        "colonne-impianti",
        Dimensions(700, 2282, 665),
        "COLONNA CON",
    ),
    (
        "colonna-libera-alto",
        ["colonna libera alto", "colonna libera in alto", "colonna libera sopra"],
        "colonne",
        Dimensions(700, 2282, 665),
        "COLONNA CON",
    ),
    (
        "colonna-libera-basso",
        ["colonna libera basso", "colonna libera in basso", "colonna libera sotto"],
        "colonne",
        Dimensions(700, 2282, 665),
        "COLONNA CON",
    ),
    (
        "contenitore-impianti",
        ["contenitore impianti"],
        "contenitori-impianti",
        Dimensions(700, 880, 665),
        "CONTENITORE IMPIANTI",
    ),
    (
        "base-lavatrice-asciugatrice",
        ["base lavatrice asciugatrice", "baselavatriceasciugatrice"],
        "basi-lavatrice-asciugatrice",
        Dimensions(700, 880, 665),
        "BASE LAVATRICE-ASCIUGATRICE",
    ),
    (
        "portale-basi-lavatrice-asciugatrice",
        ["portale basi lavatrice asciugatrice"],
        "portali",
        Dimensions(700, 880, 64),
        "PORTALE",
    ),
    (
        "base-sottolavello",
        ["base sottolavello", "basesottolavello"],
        "basi",
        Dimensions(700, 880, 665),
        "BASE SOTTOLAVELLO",
    ),
    (
        "base-sottolavatoio",
        ["base sottolavatoio", "basesottolavatoio"],
        "basi",
        Dimensions(700, 880, 665),
        "BASE SOTTOLAVATOIO",
    ),
    (
        "colonna",
        ["colonna con", "colonnacon"],
        "colonne",
        Dimensions(700, 2282, 665),
        "COLONNA CON",
    ),
    ("base", ["base con", "basecon"], "basi", Dimensions(700, 880, 665), "BASE CON"),
]


VARIANT_RULES = [
    ("two_visible_sides", ["2 fianchi a vista", "sx e dx", "a vista sx e dx"]),
    (
        "one_visible_one_internal",
        [
            "1 fianco interno 1 fianchi a vista",
            "1 fianchi a vista 1 fianco interno",
            "1 fianco a vista 1 fianco interno",
            "1 fianco interno 1 fianco vista",
            "a vista sx",
            "a vista dx",
        ],
    ),
    ("two_internal_sides", ["2 fianchi interni", "non a vista"]),
]


def normalize_text(value: str) -> str:
    """Normalizza testo e nomi file per confronti tolleranti agli errori tipografici."""
    text = re.sub(r"vista3d|\{3d\}", " ", value, flags=re.I)
    text = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", text)
    text = re.sub(r"(?<=[A-Za-z])(?=\d)|(?<=\d)(?=[A-Za-z])", " ", text)
    text = unicodedata.normalize("NFD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = text.lower()
    text = text.replace("{3d}", " ").replace("vista3d", " ")
    text = text.replace("fiaco", "fianco").replace("i nterni", "interni")
    text = text.replace("avista", "a vista").replace("nonavista", "non a vista")
    text = re.sub(r"(\d)(fianco|fianchi)", r"\1 \2", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def slugify(value: str) -> str:
    """Converte un nome in codice stabile per database e storage."""
    normalized = normalize_text(value)
    return re.sub(r"[^a-z0-9]+", "-", normalized).strip("-")


def parse_number(value: Any) -> int | None:
    """Estrae un intero da celle Excel numeriche o testuali quando possibile."""
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return int(round(value))
    if isinstance(value, str):
        match = re.search(r"\d+(?:[,.]\d+)?", value)
        if match:
            return int(round(float(match.group(0).replace(",", "."))))
    return None


def find_excel_file(incoming_dir: Path) -> Path:
    """Trova l'unico workbook Excel presente nello staging."""
    files = sorted((incoming_dir / "excel").glob("*.xlsx"))
    if len(files) != 1:
        raise RuntimeError(f"Atteso 1 file .xlsx in {incoming_dir / 'excel'}, trovati {len(files)}.")
    return files[0]


def extract_excel_systems(excel_path: Path) -> list[dict[str, Any]]:
    """Estrae le righe sistema dall'Excel mantenendo dimensioni e riga sorgente."""
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    worksheet = workbook["LEGENDA_GENERALE"]
    current_dimensions: Dimensions | None = None
    systems: list[dict[str, Any]] = []

    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        values = list(row)
        first_cell = values[0] if len(values) > 0 else None
        width = parse_number(values[2] if len(values) > 2 else None)
        height = parse_number(values[3] if len(values) > 3 else None)
        depth = parse_number(values[4] if len(values) > 4 else None)

        if isinstance(first_cell, str) and "dimensione sistema" in first_cell.lower():
            match = re.search(r"L\s*(\d+)X(\d+)HLX(\d+)", first_cell.replace(" ", ""), re.I)
            if match:
                current_dimensions = Dimensions(*map(int, match.groups()))
        elif width and height and depth:
            current_dimensions = Dimensions(width, height, depth)

        quantity_header = values[7] if len(values) > 7 else None
        system_name = values[8] if len(values) > 8 else None
        if isinstance(quantity_header, str) and "Q.TA" in quantity_header and isinstance(system_name, str):
            systems.append(
                {
                    "row": row_index,
                    "name": " ".join(system_name.split()),
                    "normalized_name": normalize_text(system_name),
                    "dimensions": asdict(current_dimensions) if current_dimensions else None,
                }
            )

    return systems


def classify_family(file_name: str) -> tuple[str, str, str, Dimensions, str | None]:
    """Deduce famiglia e categoria dal nome del GLB ricevuto."""
    normalized = normalize_text(file_name)
    compact = normalized.replace(" ", "")

    for family_slug, patterns, category_slug, dimensions, excel_hint in FAMILY_RULES:
        if any(pattern in normalized or pattern.replace(" ", "") in compact for pattern in patterns):
            return (
                family_slug,
                category_slug,
                CATEGORY_LABELS[category_slug],
                dimensions,
                excel_hint,
            )

    return ("sconosciuto", "basi", CATEGORY_LABELS["basi"], Dimensions(700, 880, 665), None)


def classify_variant(file_name: str) -> str:
    """Deduce la variante tecnica compatibile con il catalogo esistente."""
    normalized = normalize_text(file_name)
    for variant_key, patterns in VARIANT_RULES:
        if any(pattern in normalized for pattern in patterns):
            return variant_key
    return "two_visible_sides"


def find_excel_match(file_name: str, excel_hint: str | None, variant_key: str, systems: list[dict[str, Any]]) -> str | None:
    """Trova la riga sistema Excel più vicina alla famiglia e alla variante del GLB."""
    if not excel_hint:
        return None

    wanted_variant = {
        "two_visible_sides": "2 fianchi a vista",
        "one_visible_one_internal": "1 fianchi a vista",
        "two_internal_sides": "2 fianchi interni",
    }[variant_key]
    normalized_hint = normalize_text(excel_hint)
    normalized_variant = normalize_text(wanted_variant)

    for system in systems:
        name = system["normalized_name"]
        if normalized_hint in name and normalized_variant in name:
            return f"LEGENDA_GENERALE!I{system['row']}: {system['name']}"

    normalized_file = normalize_text(file_name)
    for system in systems:
        if system["normalized_name"] in normalized_file:
            return f"LEGENDA_GENERALE!I{system['row']}: {system['name']}"

    return None


def build_products(incoming_dir: Path, systems: list[dict[str, Any]]) -> list[CatalogProduct]:
    """Costruisce il manifest prodotti dai GLB locali e dalle regole di matching."""
    products: list[CatalogProduct] = []
    for glb_path in sorted((incoming_dir / "glb").glob("*.glb")):
        family_slug, category_slug, category_name, dimensions, excel_hint = classify_family(glb_path.stem)
        variant_key = classify_variant(glb_path.stem)
        code = slugify(glb_path.stem).upper().replace("-", "_")
        storage_path = f"models/{code.lower()}/{slugify(glb_path.name)}"
        review_flags: list[str] = []
        size_mb = glb_path.stat().st_size / 1024 / 1024

        if size_mb > MAX_MODEL_SIZE_MB:
            review_flags.append(f"GLB oltre {MAX_MODEL_SIZE_MB} MB ({size_mb:.1f} MB)")
        if family_slug == "sconosciuto":
            review_flags.append("Famiglia non riconosciuta automaticamente")
        if " sx" in normalize_text(glb_path.stem) or " dx" in normalize_text(glb_path.stem):
            review_flags.append("Verificare orientamento sinistra/destra nel viewer")
        if family_slug == "portale-basi-lavatrice-asciugatrice":
            review_flags.append("Profondita impostata da portale Excel: verificare se va trattato come prodotto configurabile")

        excel_match = find_excel_match(glb_path.stem, excel_hint, variant_key, systems)
        if not excel_match:
            review_flags.append("Nessuna riga sistema Excel associata")

        products.append(
            CatalogProduct(
                code=code,
                name_it=humanize_name(glb_path.stem),
                name_en=None,
                category_slug=category_slug,
                category_name=category_name,
                width_mm=dimensions.width_mm,
                height_mm=dimensions.height_mm,
                depth_mm=dimensions.depth_mm,
                thickness_mm=19.5,
                price=None,
                preview_image_url=None,
                model_url=None,
                technical_file_url=None,
                is_published=True,
                local_model_path=str(glb_path.relative_to(ROOT)),
                storage_model_path=storage_path,
                variant_key=variant_key,
                excel_match=excel_match,
                review_flags=review_flags,
            )
        )

    return products


def humanize_name(value: str) -> str:
    """Ripulisce il nome file mantenendo un'etichetta leggibile per l'admin."""
    cleaned = re.sub(r"[-_]*Vista3D-\{3D\}", "", value, flags=re.I)
    cleaned = re.sub(r"\s*\(\d+\)\s*$", "", cleaned)
    cleaned = cleaned.replace("_", " ").replace("!", " ")
    cleaned = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", cleaned)
    return " ".join(cleaned.split())


def write_manifest(products: list[CatalogProduct], output_dir: Path) -> None:
    """Scrive manifest JSON, CSV e SQL verificabili prima dell'import definitivo."""
    output_dir.mkdir(parents=True, exist_ok=True)
    product_dicts = [asdict(product) for product in products]
    (output_dir / "catalog-products.generated.json").write_text(
        json.dumps(product_dicts, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    with (output_dir / "catalog-products.generated.csv").open("w", newline="", encoding="utf-8") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=list(product_dicts[0].keys()))
        writer.writeheader()
        writer.writerows(product_dicts)

    categories = sorted({(product.category_slug, product.category_name) for product in products})
    sql_lines = [
        "-- SQL generato da tools/catalog_importer.py.",
        "-- model_url resta NULL finche i GLB non vengono caricati su Supabase Storage.",
        "begin;",
        "",
    ]
    for slug, name in categories:
        sql_lines.append(
            "insert into public.categories (name, slug, description, sort_order) "
            f"values ({sql_string(name)}, {sql_string(slug)}, null, null) "
            "on conflict (slug) do update set name = excluded.name;"
        )

    sql_lines.append("")
    for product in products:
        sql_lines.append(
            "insert into public.products (category_id, name_it, name_en, code, width_mm, height_mm, depth_mm, "
            "thickness_mm, price, preview_image_url, model_url, technical_file_url, is_published) "
            "values ("
            f"(select id from public.categories where slug = {sql_string(product.category_slug)}), "
            f"{sql_string(product.name_it)}, null, {sql_string(product.code)}, "
            f"{product.width_mm}, {product.height_mm}, {product.depth_mm}, "
            f"{product.thickness_mm}, null, null, null, null, false"
            ") on conflict (code) do update set "
            "category_id = excluded.category_id, "
            "name_it = excluded.name_it, "
            "width_mm = excluded.width_mm, "
            "height_mm = excluded.height_mm, "
            "depth_mm = excluded.depth_mm, "
            "thickness_mm = excluded.thickness_mm;"
        )
    sql_lines.extend(["", "commit;", ""])
    (output_dir / "catalog-products.generated.sql").write_text("\n".join(sql_lines), encoding="utf-8")


def sql_string(value: str) -> str:
    """Escapa una stringa per gli script SQL generati localmente."""
    return "'" + value.replace("'", "''") + "'"


def write_report(products: list[CatalogProduct], systems: list[dict[str, Any]], excel_path: Path, output_dir: Path) -> None:
    """Produce un report Markdown sintetico per la revisione manuale."""
    flagged = [product for product in products if product.review_flags]
    category_counts: dict[str, int] = {}
    variant_counts: dict[str, int] = {}
    for product in products:
        category_counts[product.category_name] = category_counts.get(product.category_name, 0) + 1
        variant_counts[product.variant_key] = variant_counts.get(product.variant_key, 0) + 1

    lines = [
        "# Audit catalogo committente",
        "",
        f"- Excel: `{excel_path.relative_to(ROOT)}`",
        f"- Sistemi Excel rilevati: {len(systems)}",
        f"- GLB/prodotti generati: {len(products)}",
        f"- Prodotti con flag da rivedere: {len(flagged)}",
        "",
        "## Conteggio categorie",
        "",
    ]
    for name, count in sorted(category_counts.items()):
        lines.append(f"- {name}: {count}")

    lines.extend(["", "## Conteggio varianti", ""])
    for variant, count in sorted(variant_counts.items()):
        lines.append(f"- `{variant}`: {count}")

    lines.extend(["", "## Flag da rivedere", ""])
    if flagged:
        for product in flagged:
            lines.append(f"- `{product.code}`: {'; '.join(product.review_flags)}")
    else:
        lines.append("- Nessun flag.")

    lines.extend(["", "## Output generati", ""])
    for file_name in [
        "catalog-products.generated.json",
        "catalog-products.generated.csv",
        "catalog-products.generated.sql",
    ]:
        lines.append(f"- `{output_dir.relative_to(ROOT) / file_name}`")

    (output_dir / "AUDIT_REPORT.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def run(incoming_dir: Path, output_dir: Path) -> None:
    """Esegue l'audit completo sul materiale caricato nello staging."""
    excel_path = find_excel_file(incoming_dir)
    systems = extract_excel_systems(excel_path)
    products = build_products(incoming_dir, systems)
    if not products:
        raise RuntimeError(f"Nessun GLB trovato in {incoming_dir / 'glb'}.")

    write_manifest(products, output_dir)
    write_report(products, systems, excel_path, output_dir)
    print(f"Excel: {excel_path.relative_to(ROOT)}")
    print(f"Sistemi Excel rilevati: {len(systems)}")
    print(f"GLB/prodotti generati: {len(products)}")
    print(f"Report: {(output_dir / 'AUDIT_REPORT.md').relative_to(ROOT)}")


def main() -> None:
    """Legge gli argomenti CLI e avvia l'audit catalogo."""
    parser = argparse.ArgumentParser(description="Audit e manifest catalogo committente.")
    parser.add_argument("--incoming-dir", type=Path, default=DEFAULT_INCOMING_DIR)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    args = parser.parse_args()
    run(args.incoming_dir, args.output_dir)


if __name__ == "__main__":
    main()
