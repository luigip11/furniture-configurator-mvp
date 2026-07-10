#!/usr/bin/env python3
"""Genera il catalogo tecnico completo dalle righe Q.TA' dell'Excel."""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from catalog_importer import (
    DEFAULT_INCOMING_DIR,
    ROOT,
    classify_family,
    classify_variant,
    find_excel_file,
    normalize_text,
)

OUTPUT_PATH = ROOT / "src" / "lib" / "configurator" / "generated-technical-catalog.ts"
MANIFEST_PATH = DEFAULT_INCOMING_DIR / "audit" / "catalog-products.generated.json"


def clean_value(value: Any):
    """Preserva numeri decimali e formule testuali eliminando celle vuote."""
    if value is None:
        return None
    if isinstance(value, str):
        text = " ".join(value.split())
        return text or None
    return value


def get_variant_from_system_name(name: str):
    """Deduce la variante tecnica dal titolo sistema dell'Excel."""
    normalized = normalize_text(name)
    if "2 fianchi interni" in normalized:
        return "two_internal_sides"
    if "1 fianchi a vista" in normalized or "1 fianco a vista" in normalized:
        return "one_visible_one_internal"
    return "two_visible_sides"


def extract_systems():
    """Estrae tutti i blocchi sistema e le relative righe distinta dall'Excel."""
    excel_path = find_excel_file(DEFAULT_INCOMING_DIR)
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    worksheet = workbook["LEGENDA_GENERALE"]
    systems = []
    current = None
    optional_next = False

    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        values = list(row)
        quantity = values[7] if len(values) > 7 else None
        name = clean_value(values[8] if len(values) > 8 else None)

        if isinstance(quantity, str) and "Q.TA" in quantity and isinstance(name, str):
            current = {
                "row": row_index,
                "name": name,
                "normalized_name": normalize_text(name),
                "variant": get_variant_from_system_name(name),
                "components": [],
            }
            systems.append(current)
            optional_next = False
            continue

        if not current or not name:
            continue

        if isinstance(name, str) and normalize_text(name) == "optional":
            optional_next = True
            continue

        if quantity is None and all(clean_value(values[index] if len(values) > index else None) is None for index in range(9, 14)):
            continue

        current["components"].append(
            {
                "quantity": clean_value(quantity),
                "name": name,
                "code": clean_value(values[9] if len(values) > 9 else None) or "",
                "widthMm": clean_value(values[10] if len(values) > 10 else None),
                "heightMm": clean_value(values[11] if len(values) > 11 else None),
                "depthMm": clean_value(values[12] if len(values) > 12 else None),
                "thicknessMm": clean_value(values[13] if len(values) > 13 else None),
                "optional": optional_next,
                "excelRow": row_index,
            }
        )
        optional_next = False

    portale_components = []
    for row_index in range(70, 75):
        values = [worksheet.cell(row_index, column).value for column in range(1, 8)]
        name = clean_value(values[0])
        if not name:
            continue

        portale_components.append(
            {
                "quantity": 1 if row_index != 74 else 2,
                "name": name,
                "code": clean_value(values[1]) or "",
                "widthMm": clean_value(values[2]),
                "heightMm": clean_value(values[3]),
                "depthMm": clean_value(values[4]),
                "thicknessMm": clean_value(values[5]),
                "excelRow": row_index,
            }
        )

    if portale_components:
        systems.append(
            {
                "row": 70,
                "name": "PORTALE BASI LAVATRICE-ASCIUGATRICE",
                "normalized_name": "portale",
                "variant": "two_visible_sides",
                "components": portale_components,
            }
        )

    return systems


def find_system_for_product(product: dict[str, Any], systems: list[dict[str, Any]]):
    """Abbina un prodotto importato al blocco tecnico Excel più specifico."""
    family_slug, _, _, _, excel_hint = classify_family(product["name_it"])
    variant = product["variant_key"]

    if family_slug == "portale-basi-lavatrice-asciugatrice":
        rows = [system for system in systems if "portale" in system["normalized_name"]]
        return rows[0] if rows else None

    if not excel_hint:
        return None

    normalized_hint = normalize_text(excel_hint)
    candidates = [
        system
        for system in systems
        if normalized_hint in system["normalized_name"] and system["variant"] == variant
    ]

    return candidates[0] if candidates else None


def build_catalog():
    """Costruisce il JSON importabile da Next con distinta per codice prodotto."""
    systems = extract_systems()
    products = json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))
    catalog = {}

    for product in products:
        system = find_system_for_product(product, systems)
        if not system:
            continue

        variant = product["variant_key"]
        catalog[product["code"]] = {
            "configurableVariants": [variant],
            "bomByVariant": {
                variant: [
                    {key: value for key, value in component.items() if value is not None}
                    for component in system["components"]
                ]
            },
            "excelSource": f"LEGENDA_GENERALE!I{system['row']}",
            "excelSystemName": system["name"],
        }

    return catalog


def main():
    """Scrive il catalogo tecnico generato e riassume le righe prodotte."""
    catalog = build_catalog()
    OUTPUT_PATH.write_text(
        'import type { ModuleTechnicalDefinition } from "./module-technical-catalog.ts";\n\n'
        "// Catalogo tecnico generato dalla legenda Excel del committente.\n"
        "export const GENERATED_TECHNICAL_CATALOG: Record<string, ModuleTechnicalDefinition> = "
        + json.dumps(catalog, indent=2, ensure_ascii=False)
        + ";\n",
        encoding="utf-8",
    )
    print(f"Generated {len(catalog)} technical definitions in {OUTPUT_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
